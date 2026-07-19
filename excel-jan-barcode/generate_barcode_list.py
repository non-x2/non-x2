#!/usr/bin/env python3
"""
products.csv(商品名, JANコード)から、本物のバーコード「画像」を
埋め込んだExcel(xlsx)を作るツール。

フォント方式のバーコード(例: JANCODE-nicotan等の専用フォント+`=jan()`関数)は、
別のPCで開く/印刷するとフォント未埋め込みで崩れて読み取れなくなることがある。
このツールは画像として貼り付けるので、どのPCで開いても・印刷しても見た目が変わらない。

使い方:
    python3 generate_barcode_list.py [入力CSV] [出力XLSX]

    引数を省略した場合は products.csv → products_with_barcode.xlsx
"""

from __future__ import annotations

import csv
import io
import shutil
import sys
from collections import Counter
from pathlib import Path

import barcode
from barcode.writer import ImageWriter
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
from PIL import Image as PILImage

BASE_DIR = Path(__file__).resolve().parent

# バーコードの物理サイズ(mm)。JAN/EAN規格(モジュール幅0.33mm=100%)より
# 少し大きめにして、安いスキャナーや多少の印刷劣化でも読み取りやすくしている。
MODULE_WIDTH_MM = 0.4   # 規格100%は0.33mm。約120%相当。
MODULE_HEIGHT_MM = 18.0  # バーの高さ。斜め読み取りに強くするため規格よりやや高め。
IMAGE_DPI = 300
DISPLAY_DPI = 96  # Excel/OOXMLが画像サイズ(px)を解釈する際の基準DPI

ROW_HEIGHT_PT = 92          # バーコード画像が縦に切れないよう余裕を持たせる
NAME_COL_WIDTH = 32
JAN_COL_WIDTH = 16
BARCODE_COL_WIDTH = 30
NOTE_COL_WIDTH = 50

HEADER_FILL = PatternFill(start_color="FFDCE6F1", end_color="FFDCE6F1", fill_type="solid")
THIN_BORDER = Border(*(Side(style="thin", color="FFB7B7B7") for _ in range(4)))

# JANコードの桁数 → (バーコードのシンボル体系, チェックデジットを除いた桁数)
_LENGTH_TO_SYMBOLOGY = {
    7: ("ean8", 7),
    8: ("ean8", 7),
    12: ("ean13", 12),
    13: ("ean13", 12),
}


def build_barcode(raw: str):
    """7/8桁(JAN-8)または12/13桁(JAN-13)の文字列から、
    チェックデジット確定済みのバーコードを作る。

    戻り値: (barcode_object, 確定したコード文字列, 注記文字列)
    """
    digits = "".join(ch for ch in raw.strip() if ch.isdigit())
    spec = _LENGTH_TO_SYMBOLOGY.get(len(digits))
    if spec is None:
        raise ValueError(
            f"JANコードは7,8,12,13桁のいずれかの数字にしてください"
            f"(入力: '{raw}' / {len(digits)}桁でした)"
        )
    symbology, body_len = spec

    body = digits[:body_len]
    obj = barcode.get(symbology, body, writer=ImageWriter())
    corrected = obj.ean  # ライブラリがチェックデジットを再計算した確定コード

    note = ""
    if len(digits) == body_len + 1 and digits[body_len] != corrected[body_len]:
        note = (
            f"チェックデジットを {digits[body_len]} → {corrected[body_len]} に自動修正"
            "(元データのチェックデジットが誤っていた可能性があります)"
        )
    return obj, corrected, note


def render_barcode_png(obj, out_path_no_ext: Path) -> Path:
    options = {
        "module_width": MODULE_WIDTH_MM,
        "module_height": MODULE_HEIGHT_MM,
        "dpi": IMAGE_DPI,
        "write_text": True,
    }
    saved_path = obj.save(str(out_path_no_ext), options)
    return Path(saved_path)


def _decode_csv_bytes(raw_bytes: bytes, csv_path: Path) -> str:
    # Excelの「CSV(コンマ区切り)」形式の既定はShift-JIS(cp932)であり、UTF-8ではない。
    # UTF-8(BOM付き含む)を先に試し、ダメならcp932にフォールバックする。
    for encoding in ("utf-8-sig", "cp932"):
        try:
            return raw_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError(
        f"CSVの文字コードを判定できませんでした(UTF-8 / Shift-JISのどちらでもありません): {csv_path}"
    )


def read_products(csv_path: Path):
    text = _decode_csv_bytes(csv_path.read_bytes(), csv_path)
    reader = csv.DictReader(io.StringIO(text))
    for i, row in enumerate(reader, start=1):
        name = (row.get("商品名") or "").strip()
        raw_jan = (row.get("JANコード") or "").strip()
        if not name and not raw_jan:
            continue  # 空行はスキップ
        yield i, name, raw_jan


def main(argv: list[str]) -> int:
    input_csv = Path(argv[1]) if len(argv) > 1 else BASE_DIR / "products.csv"
    output_xlsx = Path(argv[2]) if len(argv) > 2 else BASE_DIR / "products_with_barcode.xlsx"

    if not input_csv.exists():
        print(f"[エラー] 入力ファイルが見つかりません: {input_csv}", file=sys.stderr)
        return 1

    img_dir = BASE_DIR / "_barcode_images_tmp"
    if img_dir.exists():
        shutil.rmtree(img_dir)
    img_dir.mkdir(parents=True)

    entries = []  # (name, code, png_path, note)
    had_error = False

    try:
        products = list(read_products(input_csv))
    except ValueError as e:
        print(f"[エラー] {e}", file=sys.stderr)
        return 1

    for row_no, name, raw_jan in products:
        if not name or not raw_jan:
            print(f"[スキップ] {row_no}行目: 商品名またはJANコードが空です", file=sys.stderr)
            continue
        try:
            obj, code, note = build_barcode(raw_jan)
        except ValueError as e:
            print(f"[エラー] {row_no}行目「{name}」: {e}", file=sys.stderr)
            had_error = True
            continue

        png_path = render_barcode_png(obj, img_dir / f"barcode_{row_no}")
        entries.append([name, code, png_path, note])
        if note:
            print(f"[注記] {row_no}行目「{name}」: {note}")

    if not entries:
        print("[エラー] 有効な商品データが1件もありませんでした。", file=sys.stderr)
        return 1

    # 重複JANコードの検出(データ入力ミスの早期発見のため)
    code_counts = Counter(e[1] for e in entries)
    for entry in entries:
        code = entry[1]
        if code_counts[code] > 1:
            dup_note = f"⚠️同じJANコードが他に{code_counts[code] - 1}件あります"
            entry[3] = f"{entry[3]} / {dup_note}" if entry[3] else dup_note
            print(f"[警告] JANコード {code} が重複しています(「{entry[0]}」)")

    wb = Workbook()
    ws = wb.active
    ws.title = "商品リスト"

    last_row = len(entries) + 1

    headers = ["No.", "商品名", "JANコード", "バーコード", "備考"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = NAME_COL_WIDTH
    ws.column_dimensions["C"].width = JAN_COL_WIDTH
    ws.column_dimensions["D"].width = BARCODE_COL_WIDTH
    ws.column_dimensions["E"].width = NOTE_COL_WIDTH

    for idx, (name, code, png_path, note) in enumerate(entries, start=2):
        row_cells = {
            1: idx - 1,
            2: name,
            3: code,
            5: note,
        }
        for col, value in row_cells.items():
            cell = ws.cell(row=idx, column=col, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="center" if col in (1, 3) else "left", vertical="center"
            )
        ws.cell(row=idx, column=3).number_format = "@"  # 先頭ゼロ落ち防止(文字列として保持)
        ws.cell(row=idx, column=4).border = THIN_BORDER  # バーコード画像セルにも罫線

        ws.row_dimensions[idx].height = ROW_HEIGHT_PT

        with PILImage.open(png_path) as im:
            px_w, px_h = im.size  # IMAGE_DPIでの実ピクセル数

        xl_img = XLImage(str(png_path))
        # OOXMLは画像サイズを96DPI基準のpxとして解釈するため、
        # 実寸(mm)が変わらないよう300DPI→96DPIへ換算してから指定する。
        scale = DISPLAY_DPI / IMAGE_DPI
        xl_img.width = px_w * scale
        xl_img.height = px_h * scale

        anchor_cell = f"{get_column_letter(4)}{idx}"
        ws.add_image(xl_img, anchor_cell)

    # 表としての使い勝手(見出し固定・並べ替え/絞り込み)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{last_row}"

    # 印刷設定を既定でファイルに焼き込む。
    # 「用紙に合わせて縮小」がONだとバーコードの実寸が変わり読み取れなくなるため、
    # 常に「実際のサイズ(100%)」で印刷される状態を既定にしておく。
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=False)
    ws.page_setup.fitToPage = False
    ws.page_setup.scale = 100
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = False
    ws.page_setup.fitToHeight = False
    ws.print_area = f"A1:E{last_row}"
    ws.print_options.horizontalCentered = False

    wb.save(output_xlsx)
    shutil.rmtree(img_dir, ignore_errors=True)  # 埋め込み済みなので作業用PNGは削除
    print(f"\n作成しました: {output_xlsx}")
    print(f"{len(entries)}件のバーコードを書き出しました。")

    if had_error:
        print("一部の行はエラーのためスキップしました(上記ログ参照)。", file=sys.stderr)
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
